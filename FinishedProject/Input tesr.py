from openpyxl import load_workbook

def combine(var1,var2):
    global newVar
    var1, var2= str(var1),str(var2)
    newVar = (var1+var2)
    NewVar = str(newVar)
    return NewVar

Database = load_workbook("Rocket_Database.xlsx")
sheet = Database.active

def edit(X,Y,data):
    posistion = combine(X,Y)
    sheet[posistion] = str(data)

data1 ='a'
data2 ='1'
data3 = 'hello'
edit(data1,data2,data3)

for row in sheet:
    for cell in row:
        print(cell.value)

print(Database.sheetnames)

Database.save("Rocket_Database.xlsx")
